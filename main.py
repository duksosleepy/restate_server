import asyncio
import os
import re
from datetime import timedelta
from typing import Any, Dict, Optional

import httpx
import pandas as pd
from dotenv import load_dotenv
from pydantic import BaseModel
from restate import (
    Context,
    ObjectContext,
    RunOptions,
    Service,
    VirtualObject,
    app,
)
from restate.exceptions import TerminalError

# Load environment variables
load_dotenv()


# Data models
class HttpRequest(BaseModel):
    url: str
    data: Dict[str, Any]
    task_id: str


class HttpResponse(BaseModel):
    task_id: str
    url: str
    status_code: int
    response_data: Dict[str, Any]
    success: bool
    error: Optional[str] = None
    needs_manual_retry: bool = False


class TaskStatus(BaseModel):
    task_id: str
    detail: str  # Changed from "status" to "detail"
    response: Optional[HttpResponse] = None
    retry_count: int = 0


# Virtual Object for managing individual HTTP request tasks
http_task = VirtualObject("HttpTask")

non_existing_codes = []


@http_task.handler()
async def execute_request(
    ctx: ObjectContext, request: HttpRequest
) -> HttpResponse:
    """Execute a single HTTP request with error handling and auto-retry"""

    # Get current task status with proper type hint
    task_status = await ctx.get(
        "detail", type_hint=TaskStatus
    ) or TaskStatus(  # Changed key from "status" to "detail"
        task_id=request.task_id, detail="pending", retry_count=0
    )

    # Update status to running
    task_status.detail = "running"
    ctx.set("detail", task_status)  # Changed key from "status" to "detail"

    async def make_http_request():
        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    request.url, json=request.data, timeout=30.0
                )

                # Check for error status codes that need retry
                if response.status_code in [
                    400,
                    401,
                    403,
                    404,
                    500,
                    502,
                    503,
                    504,
                ]:
                    return HttpResponse(
                        task_id=request.task_id,
                        url=request.url,
                        status_code=response.status_code,
                        response_data=response.json()
                        if response.content
                        else {},
                        success=False,
                        error=f"HTTP {response.status_code} error",
                        needs_manual_retry=True,
                    )

                return HttpResponse(
                    task_id=request.task_id,
                    url=request.url,
                    status_code=response.status_code,
                    response_data=response.json(),
                    success=True,
                )

        except Exception as e:
            return HttpResponse(
                task_id=request.task_id,
                url=request.url,
                status_code=0,
                response_data={},
                success=False,
                error=str(e),
                needs_manual_retry=True,
            )

    # Execute the HTTP request durably
    response = await ctx.run_typed(
        "http_request", make_http_request, RunOptions(type_hint=HttpResponse)
    )

    # Store API error and extract product codes only when task fails and needs retry
    if (
        response.response_data
        and response.response_data.get("errorCode")
        and response.needs_manual_retry
    ):
        error_code = response.response_data.get("errorCode")

        # Extract and append product codes to list only when task fails and needs retry
        if isinstance(error_code, str):
            pattern = r"Mã hàng\s+([^\s]+(?:\s+[^\s]+)*?)\s+không tồn tại trong hệ thống"
            matches = re.findall(pattern, error_code)

            # Add only unique codes to prevent duplicates
            for code in matches:
                if code not in non_existing_codes:
                    non_existing_codes.append(code)

        # Ensure proper UTF-8 encoding
        if isinstance(error_code, str):
            # Decode any escaped Unicode sequences
            try:
                error_code = (
                    error_code.encode("utf-8")
                    .decode("unicode_escape")
                    .encode("latin1")
                    .decode("utf-8")
                )
            except (UnicodeDecodeError, UnicodeEncodeError):
                # If decoding fails, use the original string
                pass
        ctx.set("status", error_code)
    elif response.success:
        # Clear any previous error details on success
        ctx.clear("status")  # Changed from "api_error" to "status"

    # Handle response based on success/failure
    if response.needs_manual_retry:
        # Increment retry count
        task_status.retry_count += 1
        task_status.detail = "needs_retry"
        task_status.response = response
        ctx.set("detail", task_status)  # Changed key from "status" to "detail"

        # Schedule automatic retry after delay (exponential backoff)
        retry_delay = min(
            300, 5 * (2 ** min(task_status.retry_count, 6))
        )  # Max 5 minutes

        # Schedule retry using delayed self-invocation
        ctx.object_send(
            execute_request,
            ctx.key(),
            request,
            send_delay=timedelta(seconds=retry_delay),
        )

        return response
    else:
        # Task completed successfully - clean up state
        ctx.clear_all()  # Remove all state for this task
        return response


@http_task.handler()
async def retry_task(ctx: ObjectContext, request: HttpRequest) -> HttpResponse:
    """Manually retry a failed task"""

    # Get current task status with proper type hint
    task_status = await ctx.get(
        "detail", type_hint=TaskStatus
    )  # Changed key from "status" to "detail"
    if not task_status:
        raise TerminalError(f"Task {request.task_id} not found")

    # Increment retry count
    task_status.retry_count += 1
    task_status.detail = "running"
    ctx.set("detail", task_status)  # Changed key from "status" to "detail"

    # Re-execute the request
    return await execute_request(ctx, request)


@http_task.handler()
async def get_task_status(ctx: ObjectContext) -> Optional[TaskStatus]:
    """Get the current status of a task"""
    return await ctx.get(
        "detail", type_hint=TaskStatus
    )  # Changed key from "status" to "detail"


@http_task.handler()
async def get_status(ctx: ObjectContext) -> Optional[str]:
    """Get the API error status (errorCode) for a task"""
    return await ctx.get("status")  # This now gets the errorCode string


# Service for batch management
batch_service = Service("BatchService")

# Global variables for email timing
email_scheduler_active = False
first_submit_time = None


def generate_excel_file(codes: list) -> str:
    """Generate Excel file with non-existing codes and return filename"""
    from datetime import datetime

    from openpyxl.styles import Border, Font, PatternFill, Side

    # Create DataFrame with codes
    df = pd.DataFrame(
        {
            "Product Code": codes,
            "Status": ["Not Found"] * len(codes),
            "Detected At": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            * len(codes),
            "Action Required": ["Verify & Add to System"] * len(codes),
        }
    )

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"non_existing_codes_{timestamp}.xlsx"

    # Create Excel file with formatting
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Non-Existing Codes", index=False)

        # Get workbook and worksheet objects for openpyxl
        workbook = writer.book
        worksheet = writer.sheets["Non-Existing Codes"]

        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="D7E4BC", end_color="D7E4BC", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Format header row
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Format data cells and add borders
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border

        # Adjust column widths
        worksheet.column_dimensions["A"].width = 15  # Product Code
        worksheet.column_dimensions["B"].width = 12  # Status
        worksheet.column_dimensions["C"].width = 20  # Detected At
        worksheet.column_dimensions["D"].width = 25  # Action Required

    return filename


async def send_non_existing_codes_email(codes: list) -> None:
    """Send email with the list of existing codes and Excel attachment"""
    try:
        import smtplib
        from datetime import datetime
        from email import encoders
        from email.mime.base import MIMEBase
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        # Email configuration from environment variables
        smtp_server = os.getenv("SMTP_SERVER")
        smtp_port = int(os.getenv("SMTP_PORT"))
        email_address = os.getenv("EMAIL_ADDRESS")
        email_password = os.getenv("EMAIL_PASSWORD")
        recipients = os.getenv(
            "EMAIL_RECIPIENTS", "nam.nguyen@lug.vn,songkhoi123@gmail.com"
        )

        # Generate Excel file
        excel_filename = generate_excel_file(codes)

        # Create message
        msg = MIMEMultipart()
        msg["From"] = email_address
        msg["To"] = recipients
        msg["Subject"] = (
            f"MÃ ĐƠN HÀNG CÒN THIẾU - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )

        # Create email body
        if codes:
            body = f"""
                Thời gian xử lý: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

                Tổng số mã hàng không tồn tại trong hệ thống: {len(codes)}

                Chi tiết danh sách mã hàng được đính kèm trong file Excel.

                Vui lòng kiểm tra file đính kèm để xem danh sách đầy đủ.
                            """
        msg.attach(MIMEText(body, "plain"))

        # Attach Excel file
        with open(excel_filename, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {excel_filename}",
        )
        msg.attach(part)

        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(email_address, email_password)
        server.sendmail(
            email_address,
            recipients.split(","),
            msg.as_string(),
        )
        server.quit()

        non_existing_codes.clear()

        # Clean up Excel file after sending
        try:
            os.remove(excel_filename)
        except FileNotFoundError:
            pass

        print(
            f"Email sent successfully with {len(codes)} existing codes and Excel attachment"
        )

    except Exception as e:
        print(f"Failed to send email: {e}")
        # Clean up Excel file if email fails
        try:
            excel_filename = f"non_existing_codes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            if os.path.exists(excel_filename):
                os.remove(excel_filename)
        except:
            pass


@batch_service.handler()
async def submit_batch(ctx: Context, requests: list) -> Dict[str, str]:
    """Submit a batch of HTTP requests as individual tasks"""
    global email_scheduler_active, first_submit_time

    # Set first submit time and start email scheduler if not already active
    if not email_scheduler_active:
        first_submit_time = asyncio.get_event_loop().time()
        email_scheduler_active = True

        # Schedule single email after 10 minutes
        ctx.service_send(
            send_single_email, {}, send_delay=timedelta(minutes=10)
        )

    task_ids = []
    for req_data in requests:
        # Use Ma_Don_Hang from the request data as task ID
        # Extract from the nested structure: data.data[0].master.maDonHang
        try:
            task_id = req_data["data"]["data"][0]["master"]["maDonHang"]
        except (KeyError, IndexError, TypeError):
            # If Ma_Don_Hang is not available, generate a fallback ID
            # You might want to handle this differently based on your requirements
            raise ValueError("Ma_Don_Hang is required for task identification")

        request = HttpRequest(
            url=req_data["url"], data=req_data["data"], task_id=task_id
        )

        # Submit each request as a separate task (fire and forget)
        ctx.object_send(execute_request, task_id, request)
        task_ids.append(task_id)

    return {"message": f"Submitted {len(task_ids)} tasks", "task_ids": task_ids}


@batch_service.handler()
async def send_single_email(ctx: Context) -> Dict[str, str]:
    """Send a single email with existing codes"""
    global non_existing_codes, email_scheduler_active

    if non_existing_codes:
        # Send email with current codes
        codes_to_send = non_existing_codes.copy()
        await send_non_existing_codes_email(codes_to_send)

        # Stop the email scheduler after sending
        email_scheduler_active = False

        return {
            "message": f"Email sent with {len(codes_to_send)} codes, scheduler stopped"
        }
    else:
        # No codes to send, stop the email scheduler
        email_scheduler_active = False
        return {"message": "No codes to send, email scheduler stopped"}


@batch_service.handler()
async def stop_email_scheduler(ctx: Context) -> Dict[str, str]:
    """Stop the email scheduler"""
    global email_scheduler_active, first_submit_time

    email_scheduler_active = False
    first_submit_time = None

    return {"message": "Email scheduler stopped"}


application = app(services=[batch_service, http_task])

if __name__ == "__main__":
    import hypercorn.asyncio
    import hypercorn.config

    config = hypercorn.config.Config()
    config.bind = ["0.0.0.0:9080"]
    config.alpn_protocols = ["http/1.1"]

    asyncio.run(hypercorn.asyncio.serve(application, config))
