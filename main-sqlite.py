import asyncio
import hashlib
import json
import logging
import os
import re
import sqlite3
from datetime import datetime, timedelta
from threading import Lock
from typing import Any, Dict, Optional

import httpx
import pandas as pd
import restate
from dotenv import load_dotenv
from pydantic import BaseModel
from restate import (
    Context,
    InvocationRetryPolicy,
    ObjectContext,
    Service,
    TerminalError,
    VirtualObject,
    app,
)

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("server.log"), logging.StreamHandler()],
)

# Create a logger for database operations
db_logger = logging.getLogger("database_operations")

# Database configuration
DB_PATH = os.getenv("DB_PATH", "orders.db")
sqlite_lock = Lock()


class DatabaseManager:
    """Thread-safe SQLite3 database manager with connection pooling"""

    def __init__(self, db_path: str, max_connections: int = 10):
        self.db_path = db_path
        self.max_connections = max_connections
        self._connections = []
        self._lock = Lock()
        self._initialize_database()

    def _get_connection(self) -> sqlite3.Connection:
        """Get a database connection from the pool or create new one"""
        with self._lock:
            if self._connections:
                return self._connections.pop()
            else:
                conn = sqlite3.connect(
                    self.db_path, timeout=30.0, check_same_thread=False
                )
                # Enable WAL mode for better concurrency
                conn.execute("PRAGMA journal_mode=WAL")
                conn.execute("PRAGMA synchronous=OFF")
                conn.execute("PRAGMA cache_size=10000")
                conn.execute("PRAGMA temp_store=MEMORY")
                return conn

    def _return_connection(self, conn: sqlite3.Connection):
        """Return a connection to the pool"""
        with self._lock:
            if len(self._connections) < self.max_connections:
                self._connections.append(conn)
            else:
                conn.close()

    def _initialize_database(self):
        """Initialize database tables if they don't exist with performance optimizations"""
        with self.execute_query() as conn:
            # Performance indexes for fast queries (safe to run - IF NOT EXISTS)
            # These indexes optimize the most common query patterns
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_orders_order_id ON orders(order_id)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_orders_status_updated ON orders(status, updated_at)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_orders_first_failure ON orders(first_failure_time)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_orders_product_code ON orders(product_code)"
            )

            conn.commit()
            db_logger.info(
                "Database initialization completed with performance indexes"
            )

    def execute_query(self):
        """Context manager for database operations"""
        return DatabaseConnection(self)


class DatabaseConnection:
    """Context manager for database connections"""

    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager
        self.conn = None

    def __enter__(self) -> sqlite3.Connection:
        self.conn = self.db_manager._get_connection()
        return self.conn

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.conn:
            if exc_type is None:
                self.conn.commit()
            else:
                self.conn.rollback()
            self.db_manager._return_connection(self.conn)


# Initialize database manager
db_manager = DatabaseManager(DB_PATH)


# Data models
class HttpRequest(BaseModel):
    url: str
    data: Dict[str, Any]
    task_id: str


class HttpResponse(BaseModel):
    task_id: str
    url: str
    status_code: int
    response_data: Dict[str, Any]
    success: bool
    error: Optional[str] = None
    needs_manual_retry: bool = False
    non_existing_codes: list[str] = []  # List of unique product codes that don't exist in system


http_task = Service(
    "HttpTask",
    invocation_retry_policy=InvocationRetryPolicy(
        initial_interval=timedelta(minutes=15),
        exponentiation_factor=2.0,
        max_interval=timedelta(hours=24),  # Cap at 24 hours between retries
        max_attempts=100,  # Allow up to 100 retry attempts
        on_max_attempts='pause',  # Pause invocation instead of killing it after max attempts
    ),
    journal_retention=timedelta(days=2),  # Retain journal for 2 days after completion
    idempotency_retention=timedelta(days=2),  # Retain idempotency keys for 2 days (matches journal retention)
)



def query_from_thread(query_func, *args):
    """Execute a database query in a thread-safe manner"""
    with sqlite_lock:
        try:
            result = query_func(*args)
            return result
        except Exception as e:
            db_logger.error(
                f"Database query failed for {query_func.__name__}: {str(e)}"
            )
            raise


def update_daily_stats_thread(
    task_completed: bool = False, task_failed: bool = False
):
    """Thread function to update daily task statistics"""
    today = datetime.now().date().isoformat()

    try:
        with db_manager.execute_query() as conn:
            # Insert or update daily stats
            if task_completed:
                conn.execute(
                    """
                    INSERT INTO daily_task_stats (stat_date, completed_tasks, failed_tasks, last_updated)
                    VALUES (?, 1, 0, ?)
                    ON CONFLICT(stat_date) DO UPDATE SET
                        completed_tasks = completed_tasks + 1,
                        last_updated = ?
                """,
                    (
                        today,
                        datetime.now().isoformat(),
                        datetime.now().isoformat(),
                    ),
                )

            if task_failed:
                conn.execute(
                    """
                    INSERT INTO daily_task_stats (stat_date, completed_tasks, failed_tasks, last_updated)
                    VALUES (?, 0, 1, ?)
                    ON CONFLICT(stat_date) DO UPDATE SET
                        failed_tasks = failed_tasks + 1,
                        last_updated = ?
                """,
                    (
                        today,
                        datetime.now().isoformat(),
                        datetime.now().isoformat(),
                    ),
                )
    except Exception as e:
        db_logger.error(f"Error updating daily stats: {str(e)}")
        raise




def insert_order_thread(request: HttpRequest, response: HttpResponse):
    """Thread function to insert new order in database
    Handles orders with multiple detail items (combined by maDonHang)"""
    db_logger.info(
        f"Starting insert_order_thread for task_id: {request.task_id}"
    )

    # Extract order data from the request
    try:
        order_data = request.data["data"][0]["master"]
        details_list = request.data["data"][0].get("detail", [])
    except (KeyError, IndexError):
        order_data = {}
        details_list = []

    # Determine source_type based on maDonHang
    ma_don_hang = order_data.get("maDonHang", "")
    source_type = "offline" if "/" in ma_don_hang else "online"

    if not details_list:
        db_logger.warning(f"No detail items found for task_id: {request.task_id}")
        return f"No details to insert for order {request.task_id}"

    try:
        with db_manager.execute_query() as conn:
            # Set first_failure_time to current time for new orders
            current_time = datetime.now().isoformat()

            # Insert each detail item as a separate order record
            inserted_count = 0
            for details in details_list:
                # Use INSERT OR REPLACE to handle duplicates
                conn.execute(
                    """
                    INSERT OR REPLACE INTO orders (
                        order_id, customer_name, phone_number, document_type, document_number,
                        department_code, order_date, province, district, ward, address,
                        product_code, product_name, imei, quantity, revenue, source_type,
                        status, error_code, first_failure_time, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                    (
                        request.task_id,  # order_id
                        order_data.get("tenKhachHang"),  # customer_name
                        order_data.get("soDienThoai"),  # phone_number
                        order_data.get("maCT"),  # document_type
                        order_data.get("soCT"),  # document_number
                        order_data.get("maBoPhan"),  # department_code
                        order_data.get("ngayCT"),  # order_date
                        order_data.get("tinhThanh"),  # province
                        order_data.get("quanHuyen"),  # district
                        order_data.get("phuongXa"),  # ward
                        order_data.get("diaChi"),  # address
                        details.get("maHang"),  # product_code
                        details.get("tenHang"),  # product_name
                        details.get("imei"),  # imei
                        details.get("soLuong"),  # quantity
                        details.get("doanhThu"),  # revenue
                        source_type,  # source_type
                        "needs_retry",  # status
                        response.response_data.get(
                            "errorCode", response.error
                        ),  # error_code
                        current_time,  # first_failure_time
                        current_time,  # updated_at
                    ),
                )
                inserted_count += 1

        db_logger.info(
            f"Completed insert_order_thread for task_id: {request.task_id}, inserted {inserted_count} detail items"
        )
        return f"Inserted new order {request.task_id} with {inserted_count} detail items to database"
    except Exception as e:
        db_logger.error(f"Error inserting order {request.task_id}: {str(e)}")
        raise


def update_order_thread(request: HttpRequest, response: HttpResponse):
    """Thread function to update existing order based on order_id, product_code, and imei
    Handles orders with multiple detail items (combined by maDonHang)"""
    db_logger.info(
        f"Starting update_order_thread for task_id: {request.task_id}"
    )

    # Extract order data from the request
    try:
        order_data = request.data["data"][0]["master"]
        details_list = request.data["data"][0].get("detail", [])
    except (KeyError, IndexError):
        order_data = {}
        details_list = []

    # Determine source_type based on maDonHang
    ma_don_hang = order_data.get("maDonHang", "")
    source_type = "offline" if "/" in ma_don_hang else "online"

    if not details_list:
        db_logger.warning(f"No detail items found for task_id: {request.task_id}")
        return f"No details to update for order {request.task_id}"

    try:
        with db_manager.execute_query() as conn:
            # Update each detail item - update existing records but preserve first_failure_time
            # Only update updated_at, not first_failure_time
            updated_count = 0
            for details in details_list:
                conn.execute(
                    """
                    UPDATE orders
                    SET
                        customer_name = ?,
                        phone_number = ?,
                        document_type = ?,
                        document_number = ?,
                        department_code = ?,
                        order_date = ?,
                        province = ?,
                        district = ?,
                        ward = ?,
                        address = ?,
                        product_code = ?,
                        product_name = ?,
                        imei = ?,
                        quantity = ?,
                        revenue = ?,
                        source_type = ?,
                        error_code = ?,
                        updated_at = ?,
                        status = 'needs_retry'
                    WHERE order_id = ? AND product_code = ? AND imei = ?
                """,
                    (
                        order_data.get("tenKhachHang"),  # customer_name
                        order_data.get("soDienThoai"),  # phone_number
                        order_data.get("maCT"),  # document_type
                        order_data.get("soCT"),  # document_number
                        order_data.get("maBoPhan"),  # department_code
                        order_data.get("ngayCT"),  # order_date
                        order_data.get("tinhThanh"),  # province
                        order_data.get("quanHuyen"),  # district
                        order_data.get("phuongXa"),  # ward
                        order_data.get("diaChi"),  # address
                        details.get("maHang"),  # product_code
                        details.get("tenHang"),  # product_name
                        details.get("imei"),  # imei
                        details.get("soLuong"),  # quantity
                        details.get("doanhThu"),  # revenue
                        source_type,  # source_type
                        response.response_data.get(
                            "errorCode", response.error
                        ),  # error_code
                        datetime.now().isoformat(),  # updated_at
                        request.task_id,  # WHERE order_id
                        details.get("maHang"),  # WHERE product_code
                        details.get("imei"),  # WHERE imei
                    ),
                )
                updated_count += 1

        db_logger.info(
            f"Completed update_order_thread for task_id: {request.task_id}, updated {updated_count} detail items"
        )
        return f"Updated existing order {request.task_id} with {updated_count} detail items in database"
    except Exception as e:
        db_logger.error(f"Error updating order {request.task_id}: {str(e)}")
        raise


def handle_order_database_operation(
    request: HttpRequest, response: HttpResponse
):
    """Determine whether to insert or update based on order_id, product_code, and imei
    Handles orders with multiple detail items (combined by maDonHang)"""
    db_logger.info(
        f"Starting handle_order_database_operation for task_id: {request.task_id}"
    )

    # Extract product_codes and imeis from request
    try:
        details_list = request.data["data"][0].get("detail", [])
        if not details_list:
            db_logger.warning(f"No detail items found for task_id: {request.task_id}")
            return insert_order_thread(request, response)
        
        # Check if ANY of the detail items exist in the database
        # If at least one exists, we'll update; otherwise insert
        with db_manager.execute_query() as conn:
            exists_count = 0
            for details in details_list:
                current_product_code = details.get("maHang")
                current_imei = details.get("imei")
                
                cursor = conn.execute(
                    """
                    SELECT COUNT(*) FROM orders
                    WHERE order_id = ? AND product_code = ? AND imei = ?
                """,
                    (request.task_id, current_product_code, current_imei),
                )
                
                if cursor.fetchone()[0] > 0:
                    exists_count += 1

            exists_with_same_identifiers = exists_count > 0

        if exists_with_same_identifiers:
            # Update existing record(s)
            db_logger.info(
                f"Found {exists_count} existing detail items for task_id: {request.task_id}, updating"
            )
            return update_order_thread(request, response)
        else:
            # Insert new record(s)
            db_logger.info(
                f"No existing detail items found for task_id: {request.task_id}, inserting"
            )
            return insert_order_thread(request, response)
    except Exception as e:
        db_logger.error(
            f"Error in handle_order_database_operation for {request.task_id}: {str(e)}"
        )
        raise


def cleanup_old_failed_orders_thread(days: int = 2):
    """Thread function to cleanup orders that failed for more than specified days"""
    db_logger.info(
        f"Starting cleanup_old_failed_orders_thread for orders older than {days} days"
    )

    try:
        with db_manager.execute_query() as conn:
            # Calculate cutoff date (2 days ago by default)
            cutoff_date = (datetime.now() - timedelta(days=days)).isoformat()

            # Get orders that will be deleted
            cursor = conn.execute(
                """
                SELECT COUNT(*) FROM orders
                WHERE status = 'needs_retry' AND updated_at < ?
            """,
                (cutoff_date,),
            )
            count = cursor.fetchone()[0]

            # Delete orders older than retention period
            conn.execute(
                """
                DELETE FROM orders
                WHERE status = 'needs_retry' AND updated_at < ?
            """,
                (cutoff_date,),
            )

        db_logger.info(
            f"Completed cleanup_old_failed_orders_thread - deleted {count} orders older than {days} days"
        )
        return f"Deleted {count} old failed orders from database"
    except Exception as e:
        db_logger.error(f"Error cleaning up old failed orders: {str(e)}")
        raise


def check_retry_window_expired(task_id: str) -> bool:
    """
    Check if the retry window (2 days) has expired for a failed order.
    Returns True if the order should stop being retried.
    """
    try:
        with db_manager.execute_query() as conn:
            cursor = conn.execute(
                """
                SELECT first_failure_time FROM orders
                WHERE order_id = ? AND status = 'needs_retry'
                LIMIT 1
            """,
                (task_id,),
            )
            result = cursor.fetchone()

            if not result or not result[0]:
                # No first_failure_time found, don't expire
                return False

            first_failure_time = datetime.fromisoformat(result[0])
            current_time = datetime.now()
            days_since_first_failure = (
                current_time - first_failure_time
            ).total_seconds() / (24 * 3600)

            if days_since_first_failure >= 2:
                db_logger.info(
                    f"Retry window expired for task_id: {task_id} (failed for {days_since_first_failure:.1f} days)"
                )
                return True

            db_logger.info(
                f"Retry window active for task_id: {task_id} ({days_since_first_failure:.1f}/2 days)"
            )
            return False
    except Exception as e:
        db_logger.error(f"Error checking retry window for {task_id}: {str(e)}")
        # On error, don't expire (safer to keep retrying)
        return False


def delete_order_thread(task_id: str, request: HttpRequest = None):
    """Thread function to delete successful order from database"""
    try:
        with db_manager.execute_query() as conn:
            # Delete from orders table
            conn.execute("DELETE FROM orders WHERE order_id = ?", (task_id,))

        return f"Deleted successful order {task_id} from database"
    except Exception as e:
        db_logger.error(f"Error deleting order {task_id}: {str(e)}")
        raise


async def log_failed_order_to_db(
    ctx: Context,
    request: HttpRequest,
    response: HttpResponse,
    db_path: str = "orders.db",
) -> None:
    """Log failed order information to SQLite database"""

    # Execute database operation durably using Restate's run_typed
    result = await ctx.run_typed(
        "log_failed_order",
        lambda: query_from_thread(
            handle_order_database_operation, request, response
        ),
    )
    return result


async def delete_successful_order_from_db(
    ctx: Context,
    task_id: str,
    request: HttpRequest = None,
    db_path: str = "orders.db",
) -> None:
    """Delete successful order from database"""

    # Execute database operation durably using Restate's run_typed
    result = await ctx.run_typed(
        "delete_successful_order",
        lambda: query_from_thread(delete_order_thread, task_id, request),
    )
    return result


def purge_batch_invocation(invocation_id: str) -> str:
    """
    Purge a batch invocation from Restate storage.
    This removes the invocation metadata and journal when all tasks succeeded.

    Args:
        invocation_id: The ID of the batch invocation to purge

    Returns:
        Status message
    """

    import httpx

    try:
        # Get Restate admin API endpoint from environment or use default
        restate_admin_url = os.getenv(
            "RESTATE_ADMIN_URL", "http://localhost:9070"
        )
        purge_url = f"{restate_admin_url}/invocations/{invocation_id}/purge"

        # Use synchronous httpx client to purge the invocation
        with httpx.Client() as client:
            response = client.patch(purge_url)
            if response.status_code in [200, 202, 204]:
                db_logger.info(
                    f"Successfully purged invocation {invocation_id}"
                )
                return f"Purged invocation {invocation_id}"
            else:
                db_logger.warning(
                    f"Failed to purge invocation {invocation_id}: {response.status_code}"
                )
                return f"Failed to purge invocation {invocation_id}: {response.status_code}"
    except Exception as e:
        db_logger.error(f"Error purging invocation {invocation_id}: {str(e)}")
        raise


@http_task.handler("execute_request")
async def execute_request(ctx: Context, request: HttpRequest) -> Dict[str, Any]:
    """Execute a single HTTP request with error handling and auto-retry

    LEGACY VERSION: This handler is kept for backward compatibility with old invocations.
    New requests should use execute_request_v2 instead.
    """

    async def make_http_request():
        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    request.url, json=request.data, timeout=30.0
                )

                # Parse response data
                response_data = response.json() if response.content else {}

                # Check for error status codes
                if response.status_code in [
                    400,
                    401,
                    403,
                    404,
                    500,
                    502,
                    503,
                    504,
                ]:
                    # Return failure response - step completes successfully
                    # but we'll process the error and retry the invocation after
                    return HttpResponse(
                        task_id=request.task_id,
                        url=request.url,
                        status_code=response.status_code,
                        response_data=response_data,
                        success=False,
                        error=f"HTTP {response.status_code} error",
                        needs_manual_retry=True,
                    ).model_dump()

                # Success case - return the response
                return HttpResponse(
                    task_id=request.task_id,
                    url=request.url,
                    status_code=response.status_code,
                    response_data=response_data,
                    success=True,
                ).model_dump()

        except Exception as e:
            # Network errors - return failure response
            return HttpResponse(
                task_id=request.task_id,
                url=request.url,
                status_code=0,
                response_data={},
                success=False,
                error=str(e),
                needs_manual_retry=True,
            ).model_dump()

    # SOLUTION: Call HTTP request directly WITHOUT ctx.run_typed()
    # This is the ONLY way to get fresh HTTP requests on every retry
    # Trade-off: We lose Restate's crash recovery for the HTTP response
    # But this is acceptable because:
    #   1. CRM API is idempotent (detects duplicates)
    #   2. Fresh requests are REQUIRED when CRM state changes
    response_dict = await make_http_request()

    # Convert dict back to HttpResponse object
    response = HttpResponse(**response_dict)

    # Handle success case
    if response.success:
        db_logger.info(f"HTTP request succeeded for task_id: {request.task_id}")
        # Return the success response to complete the invocation
        return response.model_dump()

    # Handle failure case - process business logic, then raise exception to retry invocation
    if response.needs_manual_retry:
        # Store API error and extract product codes
        # Only process if errorCode exists (empty errorCode means success)
        error_code = (
            response.response_data.get("errorCode")
            if response.response_data
            else None
        )

        if error_code:  # Only process if errorCode is not empty/None
            # Only process if error_code is a string
            if isinstance(error_code, str):
                # Ensure proper UTF-8 encoding FIRST (before pattern matching)
                # This fixes Unicode escapes like "\u00e3" → "ã" in cached journal entries
                try:
                    error_code = (
                        error_code.encode("utf-8")
                        .decode("unicode_escape")
                        .encode("latin1")
                        .decode("utf-8")
                    )
                except (UnicodeDecodeError, UnicodeEncodeError):
                    pass

                # Check for duplicate document pattern first
                duplicate_pattern = r"Chứng từ\s+.+?\s+đã nhập\."
                if re.search(duplicate_pattern, error_code):
                    # Duplicate document - treat as success
                    db_logger.info(
                        f"Duplicate document detected for task_id: {request.task_id}, "
                        f"treating as success"
                    )
                    # Return success to complete the invocation
                    return HttpResponse(
                        task_id=request.task_id,
                        url=request.url,
                        status_code=response.status_code,
                        response_data=response.response_data,
                        success=True,
                        error=f"Duplicate document detected: {error_code}",
                        needs_manual_retry=False,
                    ).model_dump()

                # Extract ALL non-existing product codes from error message
                # Pattern matches: "Mã hàng <code> không tồn tại trong hệ thống"
                # Can appear multiple times in one error message
                pattern = r"Mã hàng\s+([^\s]+(?:\s+[^\s]+)*?)\s+không tồn tại trong hệ thống"
                matches = re.findall(pattern, error_code)

                # Flag to track if this error has non-existing product codes
                has_non_existing_codes = False

                # Store extracted codes in response (will be deduplicated later in submit_batch)
                if matches:
                    # Convert to list of unique codes
                    unique_codes = list(set(matches))
                    response.non_existing_codes = unique_codes
                    has_non_existing_codes = True
                    db_logger.info(
                        f"Extracted {len(unique_codes)} unique non-existing product codes "
                        f"from error for task_id: {request.task_id}: {unique_codes}"
                    )

            # Log failed order (database operation removed)
            db_logger.info(f"Failed order logged for task_id: {request.task_id}")
        else:
            # Empty errorCode means successful HTTP request - treat as success
            db_logger.info(
                f"Empty errorCode for task_id: {request.task_id}, treating as success"
            )
            # Return success to complete the invocation
            return response.model_dump()

        # Determine if we should retry based on error type:
        # 1. Non-existing product code errors (has_non_existing_codes) → RETRY
        # 2. Server errors (5xx) → RETRY
        # 3. Other client errors (400/401/403/404 without product codes) → NO RETRY (terminal)

        # First, send codes to EmailAccumulator if we extracted any
        # This must happen BEFORE raising exception
        if has_non_existing_codes and response.non_existing_codes:
            try:
                db_logger.info(
                    f"Sending {len(response.non_existing_codes)} codes to EmailAccumulator "
                    f"before retry for task_id: {request.task_id}"
                )
                await ctx.object_call(
                    add_codes,
                    key="global",
                    arg=response.non_existing_codes,
                )
            except Exception as e:
                db_logger.error(
                    f"Failed to send codes to EmailAccumulator for task_id {request.task_id}: {str(e)}"
                )
                # Continue anyway - we'll raise the retry exception

        # Now determine retry behavior
        if has_non_existing_codes:
            # Product code error - raise exception to trigger Restate retry with backoff
            db_logger.warning(
                f"HttpTask failed with non-existing product code error for task_id: {request.task_id}, "
                f"raising exception to trigger retry with backoff"
            )
            raise Exception(
                f"Non-existing product code error: {response.error} (status_code: {response.status_code})"
            )
        elif response.status_code in [400, 401, 403, 404]:
            # Other 4xx errors (duplicate, auth, etc.) - terminal, no retry
            db_logger.warning(
                f"HttpTask failed with terminal error for task_id: {request.task_id}, "
                f"status_code: {response.status_code}, completing without retry"
            )
            # Raise TerminalError to signal Restate not to retry
            raise TerminalError(
                f"Terminal HTTP error: {response.error} (status_code: {response.status_code})",
                status_code=response.status_code,
            )
        else:
            # 5xx server errors - raise exception to trigger retry with backoff
            db_logger.warning(
                f"HttpTask failed with server error for task_id: {request.task_id}, "
                f"status_code: {response.status_code}, raising exception to trigger retry with backoff"
            )
            raise Exception(
                f"HTTP request failed: {response.error} (status_code: {response.status_code})"
            )

    # Should never reach here - all paths above either return or raise
    db_logger.error(
        f"Unexpected code path for task_id: {request.task_id}, response: {response}"
    )
    raise Exception("Unexpected error in execute_request handler")


# Virtual Object for rate-limiting HTTP requests to CRM server
# Uses sequential processing with 2-second delay between requests
request_throttler = VirtualObject("RequestThrottler")


@request_throttler.handler()
async def process_request(ctx: ObjectContext, request_data: dict) -> Dict[str, Any]:
    """
    Process a single HTTP request with 2-second delay to prevent CRM overload.

    This handler uses a Virtual Object with a single key to ensure all requests
    are processed sequentially with exactly 2 seconds delay between each request.

    Args:
        request_data: Dictionary containing url, data, and task_id

    Returns:
        Status dict with task_id
    """
    # Sleep for 2 seconds before processing the request
    # This creates a durable timer that survives crashes
    await ctx.sleep(timedelta(seconds=2))

    # Create HttpRequest object from request_data
    request = HttpRequest(
        url=request_data["url"],
        data=request_data["data"],
        task_id=request_data["task_id"]
    )

    # Execute the request using service_send (fire-and-forget)
    # This ensures the throttler doesn't block waiting for the result
    # Use execute_request_v2 for new invocations (fresh HTTP on retry)
    ctx.service_send(execute_request_v2, request, idempotency_key=request.task_id)

    db_logger.info(
        f"Throttler: processed request for task_id {request.task_id} after 2-second delay"
    )

    return {
        "status": "submitted",
        "task_id": request.task_id
    }


# Virtual Object for managing non-existing codes accumulation and debounced email
email_accumulator = VirtualObject("EmailAccumulator")


@email_accumulator.handler()
async def add_codes(ctx: ObjectContext, codes: list[str]) -> Dict[str, Any]:
    """
    Add product codes to the accumulator and schedule delayed email send.

    Only ONE email is sent per day - the first time codes are added.
    Subsequent calls accumulate codes but don't schedule new emails.

    Resets accumulator automatically at the start of each new day.

    Args:
        codes: List of non-existing product codes to add

    Returns:
        Status dict
    """
    if not codes:
        return {"status": "no_codes", "count": 0}

    # Check if we're in a new day - if so, reset the accumulator
    current_date = datetime.now().date().isoformat()
    last_date = await ctx.get("last_date")

    if last_date != current_date:
        # New day detected - reset accumulator and email sent flag
        db_logger.info(f"New day detected ({current_date}), resetting accumulator from previous day ({last_date})")
        ctx.clear("codes")
        ctx.clear("email_scheduled")
        ctx.clear("email_sent_today")  # Reset the daily email flag
        ctx.set("last_date", current_date)

    # Get current accumulated codes from durable state (stored as list)
    accumulated_codes_list = await ctx.get("codes") or []

    # Convert to set for deduplication, add new codes, convert back to list
    accumulated_codes = set(accumulated_codes_list)
    accumulated_codes.update(codes)

    # Save back to durable state as list (JSON serializable)
    ctx.set("codes", list(accumulated_codes))

    db_logger.info(
        f"Added {len(codes)} codes to accumulator for {current_date}. "
        f"Total accumulated: {len(accumulated_codes)}"
    )

    # Check if email was already sent today
    email_sent_today = await ctx.get("email_sent_today") or False

    if email_sent_today:
        # Email already sent today - just accumulate codes, don't schedule new email
        db_logger.info(
            f"Email already sent today ({current_date}), accumulating codes without scheduling new email"
        )
        return {
            "status": "accumulated_no_email",
            "accumulated_count": len(accumulated_codes),
            "date": current_date,
            "message": "Codes accumulated, but email already sent today"
        }

    # Check if email is already scheduled
    email_scheduled = await ctx.get("email_scheduled") or False

    if not email_scheduled:
        # Schedule delayed email send (20 minutes from now)
        db_logger.info("Scheduling email send in 20 minutes...")

        # Mark email as scheduled
        ctx.set("email_scheduled", True)

        # Send delayed message to ourselves to trigger email after 20 minutes
        ctx.object_send(
            send_accumulated_email,
            key=ctx.key(),
            arg=None,
            send_delay=timedelta(minutes=20),
        )

        return {
            "status": "scheduled",
            "accumulated_count": len(accumulated_codes),
            "date": current_date,
            "message": "Email scheduled for 20 minutes from now"
        }
    else:
        return {
            "status": "accumulating",
            "accumulated_count": len(accumulated_codes),
            "date": current_date,
            "message": "Codes added to existing accumulator, email already scheduled"
        }


@email_accumulator.handler()
async def send_accumulated_email(ctx: ObjectContext) -> Dict[str, Any]:
    """
    Send email with all accumulated codes and reset the accumulator.
    Called automatically after 20-minute delay.

    Note: Does NOT clear last_date - this allows proper new day detection.
    """
    # Get accumulated codes (stored as list)
    accumulated_codes = await ctx.get("codes") or []
    current_date = await ctx.get("last_date") or datetime.now().date().isoformat()

    if not accumulated_codes:
        db_logger.info("No accumulated codes to send")
        # Reset the scheduled flag
        ctx.clear("email_scheduled")
        return {"status": "no_codes", "count": 0, "date": current_date}

    # Convert to sorted list for email (already a list, just sort it)
    codes_list = sorted(accumulated_codes)

    db_logger.info(f"Sending email with {len(codes_list)} accumulated codes for date {current_date}...")

    # Send email
    email_result = await ctx.run(
        "send_email",
        lambda: send_non_existing_codes_email_sync(codes_list),
    )

    if email_result["status"] == "sent":
        db_logger.info(f"Email sent successfully with {email_result['count']} codes")

        # Mark that email was sent today to prevent scheduling another email
        ctx.set("email_sent_today", True)

        # Clear the accumulator and scheduled flag after successful send
        # Keep last_date to track which day we're on
        # Keep email_sent_today to prevent scheduling new emails today
        ctx.clear("codes")
        ctx.clear("email_scheduled")

        db_logger.info(
            f"Marked email as sent for today ({current_date}). "
            f"No more emails will be scheduled until tomorrow."
        )

        return {
            "status": "sent",
            "count": email_result["count"],
            "date": current_date,
            "message": f"Email sent with {email_result['count']} codes for {current_date}"
        }
    else:
        db_logger.error(f"Email send failed: {email_result.get('error', 'unknown')}")

        # Don't clear accumulator if email failed - will retry on next batch
        ctx.clear("email_scheduled")

        return {
            "status": "failed",
            "error": email_result.get("error", "unknown"),
            "count": 0,
            "date": current_date
        }


# Service for batch management
batch_service = restate.Service("BatchService")


def generate_excel_file(codes: list) -> str:
    """Generate Excel file with non-existing codes and return filename"""
    from openpyxl.styles import Border, Font, PatternFill, Side

    # Create DataFrame with codes
    df = pd.DataFrame(
        {
            "Product Code": codes,
            "Status": ["Not Found"] * len(codes),
            "Detected At": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            * len(codes),
            "Action Required": ["Verify & Add to System"] * len(codes),
        }
    )

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"non_existing_codes_{timestamp}.xlsx"

    # Create Excel file with formatting
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Non-Existing Codes", index=False)

        # Get workbook and worksheet objects for openpyxl
        workbook = writer.book
        worksheet = writer.sheets["Non-Existing Codes"]

        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="D7E4BC", end_color="D7E4BC", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Format header row
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Format data cells and add borders
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border

        # Adjust column widths
        worksheet.column_dimensions["A"].width = 15  # Product Code
        worksheet.column_dimensions["B"].width = 12  # Status
        worksheet.column_dimensions["C"].width = 20  # Detected At
        worksheet.column_dimensions["D"].width = 25  # Action Required

    return filename


def send_non_existing_codes_email_sync(codes: list) -> dict:
    """
    Send email with the list of non-existing codes and Excel attachment.

    Args:
        codes: List of codes to send.

    Returns:
        dict with status and count of codes sent
    """
    try:
        import smtplib
        from email import encoders
        from email.mime.base import MIMEBase
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        # If no codes provided, return early
        if not codes:
            db_logger.info("No codes to email")
            return {"status": "no_codes", "count": 0}

        # Email configuration from environment variables
        smtp_server = os.getenv("SMTP_SERVER")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        email_address = os.getenv("EMAIL_ADDRESS")
        email_password = os.getenv("EMAIL_PASSWORD")
        recipients = os.getenv(
            "EMAIL_RECIPIENTS", "nam.nguyen@lug.vn,songkhoi123@gmail.com"
        )

        # Generate Excel file
        excel_filename = generate_excel_file(codes)

        # Create message
        msg = MIMEMultipart()
        msg["From"] = email_address
        msg["To"] = recipients
        msg["Subject"] = (
            f"MÃ ĐƠN HÀNG CÒN THIẾU - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )

        # Create email body
        body = f"""
            Thời gian xử lý: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

            Tổng số mã hàng không tồn tại trong hệ thống: {len(codes)}

            Chi tiết danh sách mã hàng được đính kèm trong file Excel.

            Vui lòng kiểm tra file đính kèm để xem danh sách đầy đủ.
        """
        msg.attach(MIMEText(body, "plain"))

        # Attach Excel file
        with open(excel_filename, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {excel_filename}",
        )
        msg.attach(part)

        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(email_address, email_password)
        server.sendmail(
            email_address,
            recipients.split(","),
            msg.as_string(),
        )
        server.quit()

        # Clean up Excel file after sending
        try:
            os.remove(excel_filename)
        except FileNotFoundError:
            pass

        db_logger.info(
            f"Email sent successfully with {len(codes)} non-existing codes"
        )
        return {"status": "sent", "count": len(codes)}

    except Exception as e:
        db_logger.error(f"Failed to send email: {e}")
        # Clean up Excel file if email fails
        try:
            excel_filename = f"non_existing_codes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            if os.path.exists(excel_filename):
                os.remove(excel_filename)
        except:
            pass
        return {"status": "failed", "error": str(e), "count": 0}


@http_task.handler("execute_request_v2")
async def execute_request_v2(ctx: Context, request: HttpRequest) -> Dict[str, Any]:
    """Execute a single HTTP request with error handling and auto-retry

    V2 VERSION: Makes fresh HTTP requests on every retry (no journal caching).
    This version is used for all new invocations.
    """

    async def make_http_request():
        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    request.url, json=request.data, timeout=30.0
                )

                # Parse response data
                response_data = response.json() if response.content else {}

                # Check for error status codes
                if response.status_code in [
                    400, 401, 403, 404, 500, 502, 503, 504,
                ]:
                    return HttpResponse(
                        task_id=request.task_id,
                        url=request.url,
                        status_code=response.status_code,
                        response_data=response_data,
                        success=False,
                        error=f"HTTP {response.status_code} error",
                        needs_manual_retry=True,
                    ).model_dump()

                # Success case
                return HttpResponse(
                    task_id=request.task_id,
                    url=request.url,
                    status_code=response.status_code,
                    response_data=response_data,
                    success=True,
                ).model_dump()

        except Exception as e:
            return HttpResponse(
                task_id=request.task_id,
                url=request.url,
                status_code=0,
                response_data={},
                success=False,
                error=str(e),
                needs_manual_retry=True,
            ).model_dump()

    # Make fresh HTTP request (NO ctx.run_typed caching)
    response_dict = await make_http_request()
    response = HttpResponse(**response_dict)

    # Handle success case
    if response.success:
        db_logger.info(f"HTTP request succeeded for task_id: {request.task_id}")
        return response.model_dump()

    # Handle failure case
    if response.needs_manual_retry:
        error_code = (
            response.response_data.get("errorCode")
            if response.response_data
            else None
        )

        if error_code:
            if isinstance(error_code, str):
                # Ensure proper UTF-8 encoding FIRST (before pattern matching)
                try:
                    error_code = (
                        error_code.encode("utf-8")
                        .decode("unicode_escape")
                        .encode("latin1")
                        .decode("utf-8")
                    )
                except (UnicodeDecodeError, UnicodeEncodeError):
                    pass

                # Check for duplicate document pattern
                duplicate_pattern = r"Chứng từ\s+.+?\s+đã nhập\."
                if re.search(duplicate_pattern, error_code):
                    db_logger.info(
                        f"Duplicate document detected for task_id: {request.task_id}"
                    )
                    return HttpResponse(
                        task_id=request.task_id,
                        url=request.url,
                        status_code=response.status_code,
                        response_data=response.response_data,
                        success=True,
                        error=f"Duplicate document detected: {error_code}",
                        needs_manual_retry=False,
                    ).model_dump()

                # Extract product codes
                pattern = r"Mã hàng\s+([^\s]+(?:\s+[^\s]+)*?)\s+không tồn tại trong hệ thống"
                matches = re.findall(pattern, error_code)

                has_non_existing_codes = False
                if matches:
                    unique_codes = list(set(matches))
                    response.non_existing_codes = unique_codes
                    has_non_existing_codes = True
                    db_logger.info(
                        f"Extracted {len(unique_codes)} unique non-existing product codes "
                        f"from error for task_id: {request.task_id}: {unique_codes}"
                    )
        else:
            # Empty errorCode means success
            db_logger.info(f"Empty errorCode for task_id: {request.task_id}")
            return response.model_dump()

        # Send codes to EmailAccumulator BEFORE raising exception
        if has_non_existing_codes and response.non_existing_codes:
            try:
                await ctx.object_call(
                    add_codes,
                    key="global",
                    arg=response.non_existing_codes,
                )
            except Exception as e:
                db_logger.error(f"Failed to send codes to EmailAccumulator: {str(e)}")

        # Raise exception to trigger retry
        if has_non_existing_codes:
            raise Exception(
                f"Non-existing product code error: {response.error} (status_code: {response.status_code})"
            )
        elif response.status_code in [400, 401, 403, 404]:
            raise TerminalError(
                f"Terminal HTTP error: {response.error} (status_code: {response.status_code})",
                status_code=response.status_code,
            )
        else:
            raise Exception(
                f"HTTP request failed: {response.error} (status_code: {response.status_code})"
            )

    raise Exception("Unexpected error in execute_request_v2 handler")


@batch_service.handler("submit_batch")
async def submit_batch(ctx: Context, requests: list) -> Dict[str, str]:
    """
    Submit a batch of HTTP requests through the rate-limiting throttler.

    Each request is submitted to the RequestThrottler Virtual Object which ensures
    a 1-second delay between requests to prevent CRM server overload.

    Args:
        requests: List of HTTP request data (batch of requests to process)

    Returns:
        Status dict with submitted task IDs
    """
    task_ids = []
    # Note: Non-existing codes are sent to EmailAccumulator directly from execute_request
    # before retry exception is raised

    # Process requests - submit each to the rate-limiting throttler
    for req_data in requests:
        # Generate unique task_id using hash of the entire request data
        # This ensures each order item gets a unique ID even with same maDonHang
        try:
            # Create a deterministic hash from the request data
            request_json = json.dumps(req_data, sort_keys=True)
            task_id = hashlib.md5(request_json.encode()).hexdigest()
        except (KeyError, IndexError, TypeError) as e:
            raise ValueError(f"Failed to generate task_id from request data: {e}")

        task_ids.append(task_id)

        # Prepare request data for the throttler
        request_data = {
            "url": req_data["url"],
            "data": req_data["data"],
            "task_id": task_id
        }

        # Submit to RequestThrottler using object_send (fire-and-forget)
        # All requests use the same key "global" to ensure sequential processing
        # The throttler will add 1-second delay between each request
        ctx.object_send(
            process_request,
            key="global",  # Single key ensures all requests are serialized
            arg=request_data
        )

        db_logger.info(
            f"Submitted request to throttler for task_id {task_id}"
        )

    # Batch processing complete
    # All tasks submitted to the rate-limiting throttler
    db_logger.info(
        f"Batch processing complete - submitted {len(task_ids)} tasks to throttler "
        f"(2-second delay between requests)"
    )

    return {
        "message": f"Submitted {len(task_ids)} tasks for rate-limited processing",
        "task_ids": task_ids,
        "batch_status": "submitted",
        "status_details": "All tasks submitted to rate limiter. Requests will be sent with 2-second delay between each. Each task will retry automatically on failure with 15-minute initial backoff.",
    }






@batch_service.handler("cleanup_old_failed_orders")
async def cleanup_old_failed_orders(
    ctx: Context, days: int = 2
) -> Dict[str, str]:
    """
    Cleanup orders that have failed and are older than the retention period.
    This handler should be called periodically (e.g., once per day) to clean up
    old failed invocations that are no longer being retried.

    Args:
        days: Number of days to retain failed orders (default: 2)

    Returns:
        Status message with count of deleted orders
    """
    result = await ctx.run_typed(
        "cleanup_failed_orders",
        lambda: query_from_thread(cleanup_old_failed_orders_thread, days),
    )
    return {"message": result}


application = app(services=[batch_service, http_task, request_throttler, email_accumulator])

if __name__ == "__main__":
    import hypercorn.asyncio
    import hypercorn.config

    config = hypercorn.config.Config()
    config.bind = ["0.0.0.0:9080"]
    config.alpn_protocols = ["http/1.1"]
    # Disable Hypercorn's default logging to prevent duplicate logs
    # Our logging is already configured via logging.basicConfig()
    config.accesslog = None  # Disable access logs
    config.errorlog = None   # Disable error logs (prevents duplicate logging)

    asyncio.run(hypercorn.asyncio.serve(application, config))
