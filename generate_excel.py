#!/usr/bin/env python3
"""
Generate Excel file with non-existing product codes for tech lead review
"""

import pandas as pd
from datetime import datetime
import os

# Sample non-existing codes that would typically be collected from error responses
sample_non_existing_codes = [
    "SP001",
    "SP002", 
    "SP003",
    "SP004",
    "SP005",
    "MT001",
    "MT002",
    "HW001",
    "HW002", 
    "SW001",
    "SW002",
    "DB001",
    "API001",
    "NET001",
    "SEC001"
]

def generate_excel_report():
    """Generate Excel file with non-existing codes"""
    
    # Create DataFrame with codes
    df = pd.DataFrame({
        'Product Code': sample_non_existing_codes,
        'Status': ['Not Found'] * len(sample_non_existing_codes),
        'Detected At': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')] * len(sample_non_existing_codes),
        'Action Required': ['Verify & Add to System'] * len(sample_non_existing_codes)
    })
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'non_existing_codes_{timestamp}.xlsx'
    
    # Create Excel file with formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Non-Existing Codes', index=False)
        
        # Get workbook and worksheet objects for openpyxl
        workbook = writer.book
        worksheet = writer.sheets['Non-Existing Codes']
        
        # Apply basic formatting with openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Format data cells and add borders
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 15  # Product Code
        worksheet.column_dimensions['B'].width = 12  # Status
        worksheet.column_dimensions['C'].width = 20  # Detected At
        worksheet.column_dimensions['D'].width = 25  # Action Required
    
    print(f"Excel file generated: {filename}")
    print(f"Total non-existing codes: {len(sample_non_existing_codes)}")
    return filename

if __name__ == "__main__":
    generate_excel_report()